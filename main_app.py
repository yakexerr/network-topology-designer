# main_app.py

import sys
import json
import math
import random
from dataclasses import asdict, is_dataclass
from typing import Dict, List

import openpyxl
from PyQt6.QtGui import QAction, QKeySequence
from PyQt6.QtWidgets import QApplication, QMainWindow, QMessageBox, QFileDialog, QInputDialog
from PyQt6.QtCore import Qt

# Наши модули
from ui_main_window import Ui_MainWindow
from data_models import Node, Edge, TrafficDemand
from graph_algorithms import (prim_mst, dijkstra_all_pairs_hops,
                              calculate_edge_delays, dijkstra_max_delay_path)
from routes_dialog import RoutesDialog
from evaluation_dialog import EvaluationDialog
from load_settings_dialog import LoadSettingsDialog

class EnhancedJSONEncoder(json.JSONEncoder):
    def default(self, o):
        if is_dataclass(o):
            return asdict(o)
        return super().default(o)

class MainWindow(QMainWindow, Ui_MainWindow):
    AVAILABLE_CAPACITIES = [0, 2, 4, 8, 16, 32, 64, 128, 256, 512, 1024]

    def __init__(self):
        super().__init__()
        self.setupUi(self)
        self.setWindowTitle("Проектирование топологий сетей")

        # --- Инициализация переменных ---
        self.nodes: Dict[int, Node] = {}
        self.edges: List[Edge] = []
        self.routes: Dict = {}
        self.selected_node: Node | None = None
        self.selected_edge: Edge | None = None
        self.is_move_mode = False
        self.highlighted_path: List[int] = []
        self.routes_dialog = None
        self.avg_packet_size_bits = 1500 * 8
        self.high_load_threshold = 0.6  # 60%
        self.overload_threshold = 0.9  # 90%

        self.edgeCapacityComboBox.addItems([str(c) for c in self.AVAILABLE_CAPACITIES])
        # Создаем новое действие (action)
        self.actionSetPacketSize = QAction("Задать размер пакета", self)
        # Добавляем его в меню "Этапы"
        self.menu_3.insertAction(self.actionEvaluateProject, self.actionSetPacketSize)

        # Добавляем разделитель для красоты
        self.menu_3.insertSeparator(self.actionEvaluateProject)
        self.actionLoadSettings = QAction("Настроить уровни загрузки", self)
        self.menu_3.insertAction(self.actionEvaluateProject, self.actionLoadSettings)

        self.connect_signals()
        self.update_info_panels()
        self.edgeCostEdit.setReadOnly(True)
        self.update_legend()



    def connect_signals(self):
        # Меню
        self.actionLoadFromExcel.triggered.connect(self.load_from_excel)
        self.actionLoadFromJson.triggered.connect(self.load_from_json)
        self.actionSaveAsJson.triggered.connect(self.save_as_json)
        self.actionCalculateRoutes.triggered.connect(self.calculate_routes)
        self.actionCalculateFlows.triggered.connect(self.load_traffic_and_calculate_flows)
        self.actionEvaluateProject.triggered.connect(self.evaluate_project)

        # Кнопки и чекбоксы
        self.addNodeButton.clicked.connect(self.add_node)
        self.addNodeButton.setToolTip("Добавить новый узел в центр экрана")
        self.moveModeCheckBox.stateChanged.connect(self.move_mode_changed)

        # Панели свойств
        self.nodeNameEdit.editingFinished.connect(self.update_node_properties)
        self.nodeCostEdit.editingFinished.connect(self.update_node_properties)
        self.edgeCapacityComboBox.currentIndexChanged.connect(self.manual_capacity_changed)

        # Сигналы от холста
        self.drawingCanvas.nodeSelected.connect(self.on_node_selected)
        self.drawingCanvas.edgeSelected.connect(self.on_edge_selected)
        self.drawingCanvas.selectionCleared.connect(self.on_selection_cleared)

        # Горячая клавиша Delete
        delete_action = QAction("Удалить", self)
        delete_action.setShortcut(QKeySequence.StandardKey.Delete)
        delete_action.triggered.connect(self.delete_selected_item)
        self.addAction(delete_action)

        # для ввода размера пакета
        self.actionSetPacketSize.triggered.connect(self.set_packet_size)
        self.actionLoadSettings.triggered.connect(self.open_load_settings)

    def open_load_settings(self):
        dialog = LoadSettingsDialog(self.high_load_threshold, self.overload_threshold, self)

        # exec() - модальное окно, ждет пока пользователь нажмет OK или Cancel
        if dialog.exec():
            new_values = dialog.get_values()
            self.high_load_threshold = new_values["high"]
            self.overload_threshold = new_values["overload"]

            # Обновляем легенду и холст
            self.update_legend()
            self.drawingCanvas.update()
            self.statusBar().showMessage("Уровни загрузки обновлены.", 4000)

    def update_legend(self):
        high_perc = int(self.high_load_threshold * 100)
        overload_perc = int(self.overload_threshold * 100)

        legend_html = f"""
        <b>Условные обозначения:</b><br>
        <font color='#00FF00'>■</font> - Выбранный маршрут<br>
        <font color='red'>■</font> - Выделенное ребро<br>
        <hr>
        <b>Загрузка канала:</b><br>
        <font color='#8B0000'>■</font> - Перегрузка (&ge; {overload_perc}%)<br>
        <font color='#FFA500'>■</font> - Высокая нагрузка (&ge; {high_perc}%)<br>
        <font color='black'>■</font> - Нормальная нагрузка (&lt; {high_perc}%)
        """
        self.debugOutputTextEdit.setHtml(legend_html)

        self.debugOutputTextEdit.setReadOnly(True)


    def _calculate_average_delay(self, edges: List[Edge]) -> float:
        """Рассчитывает среднюю задержку по всем загруженным каналам."""
        # 1. Собираем в список задержки только тех рёбер, где есть поток
        #    и которые не перегружены (задержка не бесконечна).
        loaded_edges_delays = [
            edge.delay for edge in edges if edge.flow > 0 and edge.delay != float('inf')
        ]
        # 2. Если таких рёбер нет, то и средней задержки нет.
        if not loaded_edges_delays:
            return 0.0
        # 3. Считаем среднее арифметическое и возвращаем.
        return sum(loaded_edges_delays) / len(loaded_edges_delays)


    def evaluate_project(self):
        if not self.edges or not any(edge.flow > 0 for edge in self.edges):
            QMessageBox.warning(self, "Ошибка", "Сначала необходимо рассчитать потоки (Этап 3).")
            return

        calculate_edge_delays(self.edges, avg_packet_size_bits=self.avg_packet_size_bits)

        # --- ИЗМЕНЯЕМ РАСЧЕТ СТОИМОСТИ ---
        total_node_cost = sum(node.cost for node in self.nodes.values())
        # Рассчитываем компоненты стоимости ребер отдельно
        total_base_edge_cost = sum(self._calculate_cost_from_length(e.length) for e in self.edges)
        total_capacity_edge_cost = sum(self._calculate_cost_from_capacity(e.capacity) for e in self.edges)

        total_project_cost = total_node_cost + total_base_edge_cost + total_capacity_edge_cost

        max_delay = dijkstra_max_delay_path(self.nodes, self.edges)
        avg_delay = self._calculate_average_delay(self.edges)

        # Передаем все компоненты в диалог
        dialog = EvaluationDialog(
            edges=self.edges,
            total_cost=total_project_cost,
            node_cost=total_node_cost,
            base_edge_cost=total_base_edge_cost,
            capacity_edge_cost=total_capacity_edge_cost,
            max_delay=max_delay,
            avg_delay=avg_delay,
            parent=self
        )
        dialog.exec()

    def save_as_json(self):
        file_name, _ = QFileDialog.getSaveFileName(self, "Сохранить проект", "", "JSON Files (*.json)")
        if not file_name: return

        try:
            data_to_save = {"nodes": list(self.nodes.values()), "edges": self.edges}
            with open(file_name, 'w', encoding='utf-8') as f:
                json.dump(data_to_save, f, cls=EnhancedJSONEncoder, indent=4, ensure_ascii=False)
            QMessageBox.information(self, "Сохранение", "Проект успешно сохранен!")
        except Exception as e:
            QMessageBox.critical(self, "Ошибка сохранения", f"Произошла ошибка:\n{e}")

    def load_from_json(self):
        file_name, _ = QFileDialog.getOpenFileName(self, "Загрузить проект", "", "JSON Files (*.json)")
        if not file_name: return

        try:
            with open(file_name, 'r', encoding='utf-8') as f:
                loaded_data = json.load(f)

            self.nodes.clear(); self.edges.clear()
            self.on_selection_cleared()

            for node_data in loaded_data["nodes"]:
                node_data['position'] = tuple(node_data['position'])
                node = Node(**node_data)
                self.nodes[node.id] = node

            for edge_data in loaded_data["edges"]:
                edge = Edge(**edge_data)
                self.edges.append(edge)

            self.drawingCanvas.update()
            QMessageBox.information(self, "Загрузка", "Проект успешно загружен!")
        except Exception as e:
            QMessageBox.critical(self, "Ошибка загрузки", f"Произошла ошибка:\n{e}")

    def _calculate_cost_from_capacity(self, capacity: float) -> float:
        """
        Вычисляет стоимость канала на основе его пропускной способности.
        Реализует ступенчатую функцию, как в твоих заметках.
        """
        # Эти "пороги" и "цены" можно вынести в настройки, если нужно
        if 0 < capacity <= 64:
            return 100.0
        elif 64 < capacity <= 128:
            return 250.0
        elif 128 < capacity <= 500:
            return 600.0
        elif capacity > 500:
            return 1000.0
        else:  # Для каналов с нулевым потоком
            return 10.0  # Минимальная стоимость за "аренду линии"

    def _calculate_cost_from_length(self, length: float) -> float:
        if 0 < length <= 100:  # По вашему графику
            return 50.0  # Условная цена
        elif 100 < length <= 300:
            return 150.0
        elif length > 300:
            return 400.0
        else:
            return 0.0


    def load_traffic_and_calculate_flows(self):
        """
        Этап 3: Загружает БЕЗЗАГОЛОВОЧНУЮ МАТРИЦУ нагрузки, рассчитывает потоки,
        подбирает пропускные способности и пересчитывает стоимость рёбер.
        """
        if not self.nodes or not self.edges:
            QMessageBox.warning(self, "Ошибка", "Сначала необходимо построить топологию.")
            return
        if not self.routes:
            QMessageBox.warning(self, "Ошибка", "Сначала необходимо рассчитать маршруты (Этап 2).")
            return

        file_name, _ = QFileDialog.getOpenFileName(self, "Выберите файл с МАТРИЦЕЙ нагрузки", "",
                                                   "Excel Files (*.xlsx)")
        if not file_name: return

        # --- Шаг 3.1: Умное чтение БЕЗЗАГОЛОВОЧНОЙ матрицы ---
        demands: List[TrafficDemand] = []
        try:
            workbook = openpyxl.load_workbook(file_name)
            sheet = workbook.active

            # --- "Умная подстановка" ---
            # Получаем отсортированный список ID узлов из нашей топологии
            sorted_node_ids = sorted(self.nodes.keys())

            # Проходим по каждой строке в Excel
            for row_index, row_cells in enumerate(sheet.iter_rows()):
                # Проверяем, что для этой строки есть соответствующий узел
                if row_index >= len(sorted_node_ids):
                    break  # Строк в Excel больше, чем у нас узлов

                from_id = sorted_node_ids[row_index]

                # Проходим по каждой ячейке в строке
                for col_index, cell in enumerate(row_cells):
                    if col_index >= len(sorted_node_ids):
                        break  # Столбцов больше, чем узлов

                    to_id = sorted_node_ids[col_index]

                    # Пропускаем диагональ (трафик от узла к самому себе)
                    if from_id == to_id:
                        continue

                    volume = cell.value
                    # Создаем требование, только если в ячейке есть число > 0
                    if volume is not None and isinstance(volume, (int, float)) and volume > 0:
                        demand = TrafficDemand(
                            from_id=from_id,
                            to_id=to_id,
                            volume=float(volume)
                        )
                        demands.append(demand)

        except Exception as e:
            QMessageBox.critical(self, "Ошибка чтения файла", f"Не удалось прочитать матрицу нагрузки:\n{e}")
            return

        print(f"Загружено и распознано {len(demands)} требований по трафику из матрицы.")

        # --- Шаги 3.2 и 3.3 ОСТАЮТСЯ АБСОЛЮТНО БЕЗ ИЗМЕНЕНИЙ! ---
        # Вся остальная логика работает с `demands` и ей неважно, как мы их получили.

        for edge in self.edges:
            edge.flow = 0.0

        for demand in demands:
            route_key = (demand.from_id, demand.to_id)
            if route_key in self.routes:
                path = self.routes[route_key]
                for i in range(len(path) - 1):
                    u, v = path[i], path[i + 1]
                    for edge in self.edges:
                        if (edge.from_id == u and edge.to_id == v) or (edge.from_id == v and edge.to_id == u):
                            edge.flow += demand.volume
                            break
            else:
                print(f"Внимание: Маршрут для {demand.from_id}->{demand.to_id} не найден.")

        for edge in self.edges:
            required_flow = edge.flow
            selected_capacity = next((c for c in self.AVAILABLE_CAPACITIES if c >= required_flow),
                                     self.AVAILABLE_CAPACITIES[-1])
            if required_flow == 0:
                selected_capacity = 0
            edge.capacity = selected_capacity

            # Стало: Считаем обе части стоимости и складываем их
            base_cost = self._calculate_cost_from_length(edge.length)
            capacity_cost = self._calculate_cost_from_capacity(edge.capacity)
            edge.cost = base_cost + capacity_cost

        print("Расчет потоков, подбор пропускных способностей и пересчет стоимостей завершен.")
        self.drawingCanvas.update()
        self.update_info_panels()
        QMessageBox.information(self, "Расчет завершен", "Потоки и пропускные способности успешно рассчитаны.")

    def calculate_routes(self):
        if len(self.nodes) < 2:
            QMessageBox.warning(self, "Ошибка", "Недостаточно узлов для расчета маршрутов.")
            return

        # Если окно уже открыто, просто показываем его и выходим
        if self.routes_dialog is not None:
            self.routes_dialog.show()
            self.routes_dialog.activateWindow()  # Этого достаточно, чтобы окно стало активным
            return

        print("Расчет маршрутов по числу хопов...")
        self.routes = dijkstra_all_pairs_hops(self.nodes, self.edges)

        # Создаем экземпляр окна и СОХРАНЯЕМ его в self
        self.routes_dialog = RoutesDialog(self.nodes, self.routes, self)
        # Подключаем его сигнал к нашему слоту для подсветки
        self.routes_dialog.routeSelected.connect(self.on_route_highlighted)
        # Подключаем сигнал о закрытии окна к нашему слоту для очистки
        self.routes_dialog.finished.connect(self.on_routes_dialog_closed)

        # Показываем окно НЕМОДАЛЬНО
        self.routes_dialog.show()

    def on_route_highlighted(self, path):
        self.highlighted_path = path
        self.drawingCanvas.update()

    # --- НОВЫЙ МЕТОД: Срабатывает при закрытии окна маршрутов ---
    def on_routes_dialog_closed(self):
        print("Окно маршрутов закрыто, сбрасываем состояние.")
        self.highlighted_path = []
        self.routes_dialog = None # "Забываем" окно, чтобы в след. раз создать его заново
        self.drawingCanvas.update()

    # --- Методы-обработчики ---

    def on_node_selected(self, node_id):
        self.selected_node = self.nodes.get(node_id)
        self.selected_edge = None
        self.update_info_panels()

    def on_edge_selected(self, edge):
        self.selected_edge = edge
        self.selected_node = None
        self.update_info_panels()

    def on_selection_cleared(self):
        self.selected_node = None
        self.selected_edge = None
        self.update_info_panels()

    def update_info_panels(self):
        """Обновляет все панели свойств на основе текущего выделения."""
        if self.selected_node:
            node = self.selected_node
            self.nodeIdEdit.setText(str(node.id))
            self.nodeNameEdit.setText(node.name)
            self.nodeXEdit.setText(str(node.position[0]))
            self.nodeYEdit.setText(str(node.position[1]))
            self.nodeCostEdit.setText(str(node.cost))
        else:
            self.nodeIdEdit.clear()
            self.nodeNameEdit.clear()
            self.nodeXEdit.clear()
            self.nodeYEdit.clear()
            self.nodeCostEdit.clear()

        if self.selected_edge:
            edge = self.selected_edge
            self.edgeNameEdit.setText(f"{edge.from_id} - {edge.to_id}")

            capacity_str = str(edge.capacity)
            # Ищем индекс этого текста в списке
            index = self.edgeCapacityComboBox.findText(capacity_str)
            if index != -1:
                # Временно отключаем сигналы, чтобы не вызвать `manual_capacity_changed`
                self.edgeCapacityComboBox.blockSignals(True)
                # Устанавливаем нужный элемент
                self.edgeCapacityComboBox.setCurrentIndex(index)
                # Включаем сигналы обратно
                self.edgeCapacityComboBox.blockSignals(False)
            self.edgeLengthEdit.setText(f"{edge.length:.2f}")
            self.edgeCostEdit.setText(f"{edge.cost:.2f}")
            self.edgeDelayEdit.setText(f"{edge.flow:.2f}")
            # Рассчитываем загрузку
            utilization = (edge.flow / edge.capacity * 100) if edge.capacity > 0 else 0
            # Устанавливаем текст в новое поле
            self.edgeUtilizationEdit.setText(f"{utilization:.2f} %")

            # --- НАШ НОВЫЙ КОД ---
            # Просто берем уже рассчитанное значение задержки из объекта ребра
            # (если оно еще не рассчитано, там будет 0.0)
            if edge.delay == float('inf'):
                delay_text = "∞ (Перегрузка)"
            else:
                delay_text = f"{edge.delay:.4f}"

            self.edgeCalculatedDelayEdit.setText(delay_text)
            # --- КОНЕЦ НОВОГО КОДА ---
        else:
            self.edgeNameEdit.clear()
            self.edgeCapacityComboBox.setCurrentIndex(0)
            self.edgeLengthEdit.clear()
            self.edgeCostEdit.clear()
            self.edgeDelayEdit.clear()
            # --- НЕ ЗАБЫВАЕМ ОЧИЩАТЬ НОВОЕ ПОЛЕ ---
            self.edgeUtilizationEdit.clear()
            self.edgeUtilizationEdit.clear()
            self.edgeCalculatedDelayEdit.clear()


    def manual_capacity_changed(self, index):
        # Если ничего не выбрано, выходим
        if not self.selected_edge or index == -1:
            return

        # 1. Запоминаем старое значение (для отката) и получаем новое
        old_capacity = self.selected_edge.capacity
        new_capacity = float(self.edgeCapacityComboBox.currentText())

        # 2. --- ВОТ ОНА, САМА ЗАЩИТА! ---
        #    Проверяем, что новое значение НЕ меньше потока.
        #    (Исключаем случай, когда мы ставим 0, это разрешено)
        if new_capacity < self.selected_edge.flow and new_capacity != 0:
            # 3. Если защита сработала - показываем пользователю предупреждение.
            QMessageBox.warning(self, "Недопустимое значение",
                                f"Нельзя установить пропускную способность ({new_capacity}) "
                                f"ниже рассчитанного потока ({self.selected_edge.flow:.2f} Мбит/с).")

            # 4. --- ОТКАТ ИЗМЕНЕНИЙ ---
            #    Возвращаем в выпадающем списке старое, правильное значение.
            #    Это именно тот баг, который мы с вами исправляли (str(int(...))).
            old_index = self.edgeCapacityComboBox.findText(str(int(old_capacity)))
            self.edgeCapacityComboBox.blockSignals(True)
            self.edgeCapacityComboBox.setCurrentIndex(old_index if old_index != -1 else 0)
            self.edgeCapacityComboBox.blockSignals(False)

            # И немедленно выходим из функции, не применяя изменений
            return

        # 5. Если защита НЕ сработала, то мы спокойно применяем новое значение
        #    и пересчитываем стоимость.
        self.selected_edge.capacity = new_capacity
        base_cost = self._calculate_cost_from_length(self.selected_edge.length)
        capacity_cost = self._calculate_cost_from_capacity(new_capacity)
        self.selected_edge.cost = base_cost + capacity_cost

        # ... и обновляем интерфейс
        self.drawingCanvas.update()
        self.update_info_panels()

    def _calculate_distance(self, p1, p2):
        return math.sqrt((p2[0] - p1[0]) ** 2 + (p2[1] - p1[1]) ** 2)

    def create_edge(self, start_node_id, end_node_id):
        if any((e.from_id == start_node_id and e.to_id == end_node_id) or \
               (e.from_id == end_node_id and e.to_id == start_node_id) for e in self.edges):
            return

        p1 = self.nodes[start_node_id].position
        p2 = self.nodes[end_node_id].position
        length = self._calculate_distance(p1, p2)

        # Ваша логика расчета стоимости из C#
        cost = self._calculate_cost_from_length(length)

        # Начальная capacity может быть любой, например 0
        new_edge = Edge(from_id=start_node_id, to_id=end_node_id, capacity=0.0, length=length, cost=cost)
        self.edges.append(new_edge)

    def delete_selected_item(self):
        print("Действие: Удалить выбранный элемент")
        if self.selected_node:
            node_id_to_delete = self.selected_node.id
            del self.nodes[node_id_to_delete]
            self.edges = [e for e in self.edges if e.from_id != node_id_to_delete and e.to_id != node_id_to_delete]
            self.on_selection_cleared()
        elif self.selected_edge:
            self.edges.remove(self.selected_edge)
            self.on_selection_cleared()

        self.drawingCanvas.update()

    def load_from_excel(self):
        file_name, _ = QFileDialog.getOpenFileName(self, "Загрузить узлы из Excel", "", "Excel Files (*.xlsx)")
        if not file_name: return

        try:
            # --- Шаг 1: Загрузка узлов (код остается прежним) ---
            self.nodes.clear();
            self.edges.clear()
            workbook = openpyxl.load_workbook(file_name)
            sheet = workbook.active
            for row in sheet.iter_rows(min_row=2):
                node = Node(id=int(row[0].value), name=str(row[1].value),
                            position=(int(row[2].value), int(row[3].value)), cost=float(row[4].value))
                self.nodes[node.id] = node

            if len(self.nodes) < 2:
                self.drawingCanvas.update()
                return

            # --- Шаг 2 (НОВЫЙ): Построение MST с помощью алгоритма Прима ---
            print("Построение Минимального остовного дерева...")
            mst_edge_tuples = prim_mst(self.nodes)
            for from_id, to_id in mst_edge_tuples:
                self.create_edge(from_id, to_id)  # Используем наш метод для создания рёбер

            # --- Шаг 3 (НОВЫЙ): Обеспечение двусвязности ---
            print("Обеспечение двусвязности...")
            node_degrees = {node_id: 0 for node_id in self.nodes}
            for edge in self.edges:
                node_degrees[edge.from_id] += 1
                node_degrees[edge.to_id] += 1

            leaf_nodes_ids = [node_id for node_id, degree in node_degrees.items() if degree == 1]

            for leaf_id in leaf_nodes_ids:
                # Находим соседа этого "листа"
                connected_neighbor = next(e.to_id if e.from_id == leaf_id else e.from_id for e in self.edges if
                                          leaf_id in (e.from_id, e.to_id))

                # Ищем кандидатов для новой связи
                candidates = [
                    node_id for node_id in self.nodes
                    if node_id != leaf_id and node_id != connected_neighbor
                ]
                # Убираем тех, с кем уже есть связь
                candidates = [c for c in candidates if not any(
                    (e.from_id == leaf_id and e.to_id == c) or (e.from_id == c and e.to_id == leaf_id) for e in
                    self.edges)]

                if candidates:
                    target_id = random.choice(candidates)
                    print(f"Добавляем резервное ребро от {leaf_id} к {target_id}")
                    self.create_edge(leaf_id, target_id)

            self.drawingCanvas.update()
            QMessageBox.information(self, "Загрузка завершена",
                                    f"Успешно загружено {len(self.nodes)} узлов. Топология построена.")
        except Exception as e:
            QMessageBox.critical(self, "Ошибка чтения файла", f"Не удалось прочитать файл Excel:\n{e}")

    def add_node(self):
        print("Действие: Добавить узел")
        new_id = max(self.nodes.keys()) + 1 if self.nodes else 0
        pos_x = self.drawingCanvas.width() // 2
        pos_y = self.drawingCanvas.height() // 2
        new_node = Node(id=new_id, position=(pos_x, pos_y), name=f"Node{new_id}", cost=0.0)
        self.nodes[new_id] = new_node
        self.drawingCanvas.update()

    def move_mode_changed(self, state):
        self.is_move_mode = (state == Qt.CheckState.Checked.value)
        print(f"Действие: Режим перемещения изменен на {self.is_move_mode}")

    def update_node_properties(self):
        if not self.selected_node: return
        print(f"Действие: Обновить свойства узла {self.selected_node.id}")
        self.selected_node.name = self.nodeNameEdit.text()
        try:
            self.selected_node.cost = float(self.nodeCostEdit.text())

        except ValueError:
            QMessageBox.warning(self, "Ошибка ввода", "Стоимость и производительность должны быть числами.")
            self.nodeCostEdit.setText(str(self.selected_node.cost))
        self.drawingCanvas.update()

    def update_edge_properties(self):
        if not self.selected_edge: return
        print(f"Действие: Обновить свойства ребра {self.selected_edge.from_id}-{self.selected_edge.to_id}")
        try:
            self.selected_edge.capacity = float(self.edgeCapacityEdit.text())
        except ValueError:
            QMessageBox.warning(self, "Ошибка ввода", "Пропускная способность должна быть числом.")
            self.edgeCapacityEdit.setText(str(self.selected_edge.capacity))
        self.drawingCanvas.update()

    def set_packet_size(self):
        # Переводим биты обратно в байты для удобства пользователя
        current_size_bytes = self.avg_packet_size_bits // 8

        new_size_bytes, ok = QInputDialog.getInt(self, "Настройка проекта",
                                                 "Введите средний размер пакета в байтах:",
                                                 value=current_size_bytes, min=64, max=9000)
        if ok:
            self.avg_packet_size_bits = new_size_bytes * 8
            self.statusBar().showMessage(f"Размер пакета установлен: {new_size_bytes} байт.", 5000)
            # Если уже были расчеты, их нужно сбросить, так как задержки изменятся!
            # Найдем ребра, у которых есть задержка, и сбросим ее
            if any(edge.delay > 0 for edge in self.edges):
                for edge in self.edges:
                    edge.delay = 0.0
                self.statusBar().showMessage(f"Размер пакета изменен. Задержки сброшены.", 5000)


if __name__ == '__main__':
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec())